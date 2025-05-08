import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import os

# File setup
file_name = "student_scores.xlsx"
if not os.path.exists(file_name):
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Score", "Result"])
    wb.save(file_name)

# Load workbook
wb = load_workbook(file_name)
ws = wb.active

# Functions
def save_score():
    name = entry_name.get()
    try:
        score = int(entry_score.get())
        result = "Pass" if score >= 75 else "Fail"
        ws.append([name, score, result])
        wb.save(file_name)
        messagebox.showinfo("Success", f"Saved: {name} - {score} - {result}")
        entry_name.delete(0, tk.END)
        entry_score.delete(0, tk.END)
    except ValueError:
        messagebox.showerror("Error", "Please enter a valid number for score.")

def show_records():
    records_window = tk.Toplevel(window)
    records_window.title("All Student Records")
    
    text = tk.Text(records_window, width=40, height=20)
    text.pack(padx=10, pady=10)

    text.insert(tk.END, "Name\tScore\tResult\n")
    text.insert(tk.END, "-"*30 + "\n")
    for row in ws.iter_rows(min_row=2, values_only=True):
        text.insert(tk.END, f"{row[0]}\t{row[1]}\t{row[2]}\n")

# GUI setup
window = tk.Tk()
window.title("Student Score Tracker")

tk.Label(window, text="Student Name:").grid(row=0, column=0, padx=5, pady=5)
entry_name = tk.Entry(window)
entry_name.grid(row=0, column=1, padx=5, pady=5)

tk.Label(window, text="Score:").grid(row=1, column=0, padx=5, pady=5)
entry_score = tk.Entry(window)
entry_score.grid(row=1, column=1, padx=5, pady=5)

button_save = tk.Button(window, text="Save Score", command=save_score)
button_save.grid(row=2, column=0, padx=5, pady=10)

button_show = tk.Button(window, text="Show Records", command=show_records)
button_show.grid(row=2, column=1, padx=5, pady=10)

window.mainloop()
